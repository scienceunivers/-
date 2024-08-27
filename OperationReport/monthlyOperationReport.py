import openpyxl,xlrd,xlwings
import pandas as pd, numpy as np
import datetime as dt

import os

# openpyxl==2.4
# xlwings==0.11

workbook_orgTable_all=openpyxl.load_workbook(r'D:\...\gxpt\cyzb\1-jgjrzb20210203.xlsx',
                                            data_only=True)
sheetNames = workbook_orgTable_all.sheetnames
print(sheetNames)
objectiveSheet = 'Sheet1'

nCols = workbook_orgTable_all[objectiveSheet].max_column
nRows = workbook_orgTable_all[objectiveSheet].max_row
print(nRows,nCols)
colNames = []
# colNames_Row1,colNames_Row2 = [],[]
workbook_orgTable_all_dict = {}
# for j in range(0,nRows):
#     if any(wworkbook_orgTable_all[objectiveSheet].iter_rows(max_col=nCols, max_row=nRows)):
#         print(j)
#         break
for jRow in workbook_orgTable_all[objectiveSheet].iter_rows(min_row=None, max_row=nRows, min_col=None, max_col=nCols, 
                                                            ):
    for jCell in jRow:
        if jCell.value:
            
            # 没用worksheet的values属性，因为values是按行返回值，
            # 但没有对偶的按列返回值的属性，如果遇到了按列返回值的需求还得重新写循环。
            print(jCell.row)
            # openpyxl计数大部分是1-base。
            jCell_value=jCell.value
            break
    if jCell.value:
        break

# 判定从哪行开始有数据，不能用workbook_orgTable_all[objectiveSheet].min_row
for iRow in workbook_orgTable_all[objectiveSheet].iter_rows(min_col=None, max_col=None, min_row=jCell.row, 
                                                                         max_row=jCell.row):
    # 迭代出表头这一行，将这一行单元格的值append到colNames中。
    for iCell in iRow:
        colNames.append(iCell.value)
        workbook_orgTable_all_dict[colNames[iCell.column-1]] = []
        # 不如xlrd直接用col_values返回list方便。
        for kColumn in workbook_orgTable_all[objectiveSheet].iter_cols(min_col=iCell.column, 
                                                                         max_col=iCell.column,
                                                                         min_row=iCell.row+1, max_row=None):
            for kCell in kColumn:
                workbook_orgTable_all_dict[colNames[iCell.column-1]].append(kCell.value)
                      
        # iCell.column-1，报错TypeError: unsupported operand type(s) for -: 'str' and 'int'，2.4版本源代码中确实把column
        # 写成了获得单元格列字母的属性，col_idx是单元格列号的属性，到2.6将column改成了列号，获得列字母的属性改成了get_column_letter。
        
# 读取工作簿的代码，也可以用Worksheet.cell(row,column).value的思路重写一下。
data = pd.DataFrame(workbook_orgTable_all_dict, columns=colNames)
originalData_jieruzongbiao=data.loc[:,:'zsjrsjc'].copy()

# openpyxl把日期列的空单元格都转化成了NaT，把文本列的空单元格转化成了NAN。



# 侦测空格：

# originalData_jieruzongbiao.where(originalData_jieruzongbiao==' ')
print(
    originalData_jieruzongbiao.loc[originalData_jieruzongbiao.isin([' ']).any(axis=1)].index,'\n',
    originalData_jieruzongbiao.loc[:,originalData_jieruzongbiao.isin([' ']).any(axis=0)].columns
)
originalData_jieruzongbiao.replace(to_replace=r'^[ \t\n\r\f\v]+|[ \t\n\r\f\v]+$',value=np.nan,regex=True,inplace=True)
# 如果to_replace, value参数中任何一个采取regex形式，需令regex=True。如果value=None，则不替换。虽然None与np.nan
# 在df中涉及df数值运算时被同等对待，但在字符串操作比如替换中，二者又不一样。

originalData_jieruzongbiao.fillna({'xysjc':pd.NaT,'pxsjc':pd.NaT,'jrcssjc':pd.NaT,'zsjrsjc':pd.NaT}
                                          ,inplace=True)
print(
    originalData_jieruzongbiao.loc[originalData_jieruzongbiao.isin([' ']).any(axis=1)].index,'\n',
    originalData_jieruzongbiao.loc[:,originalData_jieruzongbiao.isin([' ']).any(axis=0)].columns
)
originalData_jieruzongbiao.loc[:,['xysjc','pxsjc','jrcssjc','zsjrsjc']].dtypes 
# 验证各类时间戳类型。
# originalData_jieruzongbiao['zsjrsjc'].apply(lambda x: 0 if isinstance(x,(dt.datetime,pd._libs.tslibs.nattype.NaTType)) else 1)
# 后续如有其它脏数据，再改regex。



# 后续完善数据质量检测代码：
# excel把空格也当做筛选中的空白。
# 虽然python认为None ==None为真，但在pandas里，None按np.nan对待，不建议在pandas中用None ==None。



officiallyParticiOrg_jieruzongbiao=originalData_jieruzongbiao.loc[
    (originalData_jieruzongbiao['hzxy']!='4xyjc')
        &~(originalData_jieruzongbiao['zsjr'].isna())
           &(originalData_jieruzongbiao['zsjrsjc']<=dt.datetime(2021,1,31)),:
    ]
        
officiallyParticiOrg_jieruzongbiao_danweiquancheng=officiallyParticiOrg_jieruzongbiao.loc[:,'单位全称'
    ]
print(
    '截至月末机构{}家({:%Y%m%d})'.format
    (
   officiallyParticiOrg_jieruzongbiao.shape[0],dt.datetime.today()
    )
)
# openpyxl\cell.py, openpyxl\styles\numbers.py源代码，用其中方法
# openpyxl.styles.numbers.is_date_format(workbook_orgTable_all[objectiveSheet].cell(779,21).value)去判定空时间戳类型。
officiallyParticiOrg_jieruzongbiao_20210203=officiallyParticiOrg_jieruzongbiao.copy()



originalData_jierujigouhuizongbiao = pd.read_excel(
    r'...\gxpt\cyzb\202102202055gxptdqjrjgqkhz.xlsx',
    sheet_name=0,header=0, names=None, index_col=None, 
    usecols=None, squeeze=False, dtype=None, engine=None, 
    converters=None, true_values=None, false_values=None, skiprows=None, nrows=None, 
    na_values=None, keep_default_na=True, verbose=False, 
    parse_dates=False, date_parser=None, thousands=None, comment=None, 
    skipfooter=0, convert_float=True, mangle_dupe_cols=True
)
originalData_jiekoujierubiao = pd.read_excel(
    r'...\gxpt\cyzb\6-T+1jkbsjdb20210201.xlsx',
    sheet_name=0,header=0, names=None, index_col=None, 
    usecols=None, squeeze=False, dtype=None, engine=None, 
    converters=None, true_values=None, false_values=None, skiprows=None, nrows=None, 
    na_values=None, keep_default_na=True, verbose=False, 
    parse_dates=False, date_parser=None, thousands=None, comment=None, 
    skipfooter=0, convert_float=True, mangle_dupe_cols=True
)

# 1224后续要读取两个月份的jkjrjdb，还是把更新日期列统一成datetime64更方便。通过如下代码：
# originalData_jiekoujierubiao.loc[
#     originalData_jiekoujierubiao['更新日期'].str.contains(r'^[ \t\n\r\f\v]+|[ \t\n\r\f\v]+$', regex=True).notna(),:]
# 检测出某个单元格是string且不含whitespace，是False，非string的元素都返回NaN，这应该就是字符了，输出发现果然是备注列填到了更新日期列。
officiallyParticiOrg_jiekoujierubiao=originalData_jiekoujierubiao.loc[
    (originalData_jiekoujierubiao['jzqk']=='已完成')
    &(
        (originalData_jiekoujierubiao['更新日期'].isna())
        |(originalData_jiekoujierubiao['更新日期']<=dt.datetime(2021,1,31))
    ),:
]
# pd.NaT>dt.datetime(2020,11,30)，无论>, <,都返回False，与nan类似。
print(
    '截至月末机构{}家，机构{}家({:%Y%m%d})'.format
    (
        officiallyParticiOrg_jiekoujierubiao.shape[0],
        officiallyParticiOrg_jieruzongbiao.shape[0]-officiallyParticiOrg_jiekoujierubiao.shape[0],
        dt.datetime.today()
    )
)

officiallyParticiOrg_jiekoujierubiao_20210201=officiallyParticiOrg_jiekoujierubiao.copy()


# 两个表中的zsjrjg进行比较。
# 检验两个集合的机构名称或者机构社会信用代码集合是否相等。
# 有一个命题，集合A、B非空，A∪B=U，则A=B的充要条件为(A-B)=(B-A)=∅。
# 思想证明：必要性一眼就看出来，充分性，A-B与B-A是互斥的（由定义），在Venn图中不存在交集，它们要相等，只能都是空集。
# 符号证明略。
print(set(officiallyParticiOrg_jieruzongbiao_danweiquancheng)-set(officiallyParticiOrg_jierujigouhuizongbiao_jierujigouquancheng),'\n',
set(officiallyParticiOrg_jierujigouhuizongbiao_jierujigouquancheng)-set(officiallyParticiOrg_jieruzongbiao_danweiquancheng))
# 后续可以写一个prompt，输出集合运算结果，让人判断是否一致（上个月jrhzb代表了10月的jrzbzsjrjg，所以这个对比
# 也是本期11月与上期10月jrzb的对比，当然如果读取的是10月jrzb和10月jrhzb，那二者是一样的），一致则往下运行，
# 不一致则看是否要对jrhzb执行补充操作。
# 一般是jrzb比jrhzb多，将多的机构补充到jrhzb中。

officiallyParticiOrg_jierujigouhuizongbiao_huiyuanguanlianhuiyuan=officiallyParticiOrg_jierujigouhuizongbiao.loc[
     ~officiallyParticiOrg_jierujigouhuizongbiao['shifouhuiyuanhuoguanlian'].isna(),:
 ]
officiallyParticiOrg_jierujigouhuizongbiao_feihuiyuan=(officiallyParticiOrg_jieruzongbiao.shape[0]
                                                       -officiallyParticiOrg_jierujigouhuizongbiao_huiyuanguanlianhuiyuan.shape[0])
print(
    '机构{}家({:%Y%m%d})'.format
    (
        officiallyParticiOrg_jierujigouhuizongbiao_huiyuanguanlianhuiyuan.shape[0],dt.datetime.today()
    ),
    '机构{}家({:%Y%m%d})'.format
    (
        officiallyParticiOrg_jierujigouhuizongbiao_feihuiyuan,dt.datetime.today()
    ),
    sep='\n'
)
# 这里发现了print的一个特性，print 2个及以上object，如果print与它的左括号不在一行，则会输出一个tuple。
# print与它的左括号在一行则正常输出。


# 1212总结的代码原则：利用数据集约化，形成复杂数据结构，来简化代码。
# 将值归类，写成特定数据结构，让多行代码可以用循环、序列完成。
# 进一步增加数据结构复杂度，简化循环、判断的嵌套层级。
workbookName=['jrzb','cszhzb','zszhzb','cxjrzb','T+1jkbsjdb','zdzjdxxzl',
              'jrhzb']
workbookDataframe={'jrzb':originalData_jieruzongbiao,
                   'cszhzb':None,
                   'zszhzb':None,
                   'cxjrzb':None,
                   'T+1jkbsjdb':None,
                   'zdzjdxxzl':None,
                   'jrhzb':originalData_jierujigouhuizongbiao
                  }
officiallyParticiOrgDataframe={'zsjrjg':[officiallyParticiOrg_jieruzongbiao,officiallyParticiOrg_jierujigouhuizongbiao]
                   }

# 会员这里两张表中对应列的值不是一套，按原来的思路写，还得写很多判断、try catch，与其在代码中嵌套很多层，
# 不如在数据字典中多做一些集约化，或者compact design。
# 把orgTypeColomn扩充为orgTypeAndMemberTypeValueColomn是个非常好的创意，将代码复杂性转移到数据结构上。
orgTypeColomn={'jrzb':'有无牌照',
               'jrhzb':'机构类型'
                                }
orgTypeAndMemberTypeValueColomn={'jrzb':
                                 ('有无牌照','是否会员',['是']),
                                 'jrhzb':
                                 ('机构类型','shifouhuiyuanhuoguanlian',['是','guanlian'])
                                }

# memberTypeColumn={'jrzb':,'jrhzb':
#                  }
memberType=['是','是会员','guanlian']
orgType=['cpxjjg','hlwxfjr','jys','wlxd','xedk','xt','zdjg','yh']
# 机构类型能否通过openpyxl读取出来？可以啊，set就行了。

# targetMonth=9
# 为了以后方便计算任何月份的月报，就不把dt.date.today().month-1写在代码里了，统一在全局的角度赋值即可。


# 1210循环写法
for dataframe in officiallyParticiOrgDataframe['zsjrjg']:
    # 对于每张机构接入表格。
    orgCount,memberCount,relatedMemberCount=[],[],[]
    for keys,values in orgTypeAndMemberTypeValueColomn.items():
        # 对于每张表格中的机构类型列。
        # dataframe可能与values发生错配，需要try，忽略错配形成的KeyError。
        try:
            print(dataframe[values[0]].name,dataframe[values[1]].name)
            for types in orgType:
                # 对于任一表格中，与对应的机构类型列，任一业态的统计数据。
                orgCount.append(dataframe.loc[dataframe[values[0]]==types,:].shape[0])
                # 会员这里两张表中对应列的值不是一套，所以还得写很多判断、try catch，与其在代码中嵌套很多层，
                # 不如在数据字典中多做一些集约化，或者compact design。
                # 把orgTypeColomn扩充为orgTypeAndMemberTypeValueColomn是个非常好的创意，将代码复杂性转移到数据结构上。
                memberCount.append(
                    (dataframe.loc[
                        (dataframe[values[0]]==types)
                        &(dataframe[values[1]]==values[2][0]),:].shape[0])
                )
#                 print(keys,types,memberCount[-1])
                try:
                    relatedMemberCount.append(
                                        (dataframe.loc[
                            (dataframe[values[0]]==types)
                            &(dataframe[values[1]]==values[2][1]),:].shape[0])
                    )
                except IndexError:
                    pass
                print('{keys}:{types}:{num},{memberCountAsOfType},{relatedMemberCount}'.format(
                    keys=keys,types=types,num=orgCount[-1],
                    memberCountAsOfType=memberCount[-1],
                    relatedMemberCount=relatedMemberCount[-1] if len(relatedMemberCount)>0 else 'None',
                    width=30)
                     )
                # relatedMemberCount由于jrzb中没有这一项统计，可能生成[]，所以需要加一个ifelse判断，否则index out of range。
            
        except KeyError as e:
            continue
        print('{keys}:{newline}{Dict}'.format(keys=keys,newline='\n',Dict=dict(zip(orgType,orgCount))),
             '\n',
             '{keys}memberCount:{newline}{Dict}'.format(keys=keys,newline='\n',Dict=dict(zip(orgType,memberCount))),memberCount,
              '\n',
              '{keys}relatedMemberCount:{newline}{Dict}'.format(keys=keys,newline='\n',Dict=dict(zip(orgType,relatedMemberCount)))
             )
        print(orgCount,memberCount,relatedMemberCount)
        # 我曹，1223发现一个问题，jrhzb的zhudai会员数量是3，但对应的表里明明是4个，我用loc筛选一下，确实是4个zhudai会员机构。
        # 怀疑是变量传递有问题，但打印的都是memberCount，分业态打印时，打印出来的数量是4个，是对的，用字典形式打印出来，就是3
        # 个，很可能问题出在形成字典的过程中。也排除了把jrzb中的zhudai会员数量3传递到jrhzb这里的可能（即df与统计的数量
        # 错配的可能）。试着把所有有关的变量都打印了一下，发现memberCount由于一直在被append，从内层循环到最外层，没有被清空过，
        # 所以最后它的长度是业态个数*2，所以进行zip的时候会出现不匹配，进而数据丢失或错配。
        # 我在最后增加了print(orgCount,memberCount,relatedMemberCount)，打印出memberCount，证实了这个猜测。
        # 把orgCount,memberCount,relatedMemberCount=[],[],[]从循环外面拿到第一层循环里面，可以解决这个问题。
        
        
        
        originalData_jieruzongbiao_shenqingjieru=originalData_jieruzongbiao.loc[
    (~originalData_jieruzongbiao['sfpx'].isna())
    &(dt.datetime(2021,12,1)<=originalData_jieruzongbiao['pxsjc'])
    &(originalData_jieruzongbiao['pxsjc']<=dt.datetime(2021,12,31)),:
]
# originalData_jieruzongbiao_shenqingjieru=originalData_jieruzongbiao_m1.loc[
#     (dt.datetime(2020,11,1)<=originalData_jieruzongbiao_m1['pxsjc'])
#     &(originalData_jieruzongbiao_m1['pxsjc']<=dt.datetime(2020,11,30)),:
# ]

originalData_jieruzongbiao_jinrushengchan=originalData_jieruzongbiao.loc[
    (~originalData_jieruzongbiao['zsjrsjc'].isna())
    &(dt.datetime(2021,12,1)<=originalData_jieruzongbiao['zsjrsjc'])
    &(originalData_jieruzongbiao['zsjrsjc']<=dt.datetime(2021,12,31)),:
]
# originalData_jieruzongbiao_jinrushengchan=originalData_jieruzongbiao_m2.loc[
#     (dt.datetime(2020,9,1)<=originalData_jieruzongbiao_m2['zsjrsjc'])
#     &(originalData_jieruzongbiao_m2['zsjrsjc']<=dt.datetime(2020,9,30)),:
# ]
print
(
    '机构{}家，{}家({:%Y%m%d})'.format
    (
        originalData_jieruzongbiao_shenqingjieru.shape[0],
        originalData_jieruzongbiao_jinrushengchan.shape[0],
        dt.datetime.today()
    )
)


# 读取上月月报对应的jrzb。
# workbook读取代码块，已用r'D:\hlwjrxh\...\1-1各月业务量(2020-11-17).xlsx'做了测试。
workbook_orgTable_all=openpyxl.load_workbook(r'D:\hlwjrxh\...\1-jrzb20210928.xlsx',
                                            data_only=True)
# 还是加上data_only，后续需要添加公式，统一添加便了。
sheetNames = workbook_orgTable_all.sheetnames
print(sheetNames)
objectiveSheet = 'Sheet1'
# workbook_orgTable_all_data = pd.DataFrame(workbook_orgTable_all) 不能直接转化成df
nCols = workbook_orgTable_all[objectiveSheet].max_column
nRows = workbook_orgTable_all[objectiveSheet].max_row
print(nRows,nCols)
colNames = []
workbook_orgTable_all_dict = {}

for jRow in workbook_orgTable_all[objectiveSheet].iter_rows(min_row=None, max_row=nRows, min_col=None, max_col=nCols, 
                                                            ):
    # values_only=False这个参数2.6及以上才有，我目前是2.4。
    for jCell in jRow:
        if jCell.value:
            # library CH4 定义了哪些object会被视为False。
            # 我没用worksheet的values属性，因为values是按行返回值，
            # 但没有对偶的按列返回值的属性，如果遇到了按列返回值的需求还得重新写循环。
            print(jCell.row)
            # 注意openpyxl计数大部分是1-base。
            jCell_value=jCell.value
            break
    if jCell.value:
        break
    # 内层for中放2个break不管用，外层for的末尾放break会让外层for运行完第一次就提前break，得加个判断。
# 判定从哪行开始有数据，不能用workbook_orgTable_all[objectiveSheet].min_row，这个返回1，显然第一行没数据。
# 不知道min_row对有无数据是怎么判定的。
for iRow in workbook_orgTable_all[objectiveSheet].iter_rows(min_col=None, max_col=None, min_row=jCell.row, 
                                                                         max_row=jCell.row):
    # 迭代出表头这一行，将这一行单元格的值append到colNames中。
    # iter_rows很灵活的方法。
    for iCell in iRow:
        colNames.append(iCell.value)
        # 先给字典每个键值赋一个空list，再append，[].append不返回任何值，不能直接赋值给其他变量。
        workbook_orgTable_all_dict[colNames[iCell.column-1]] = []
        # 对于表头每个cell所在的列，把这一列的值append到字典中。不能用list comprehension，否则把嵌套list赋给字典了。
        # 还是得套循环，这不如xlrd直接用col_values返回list方便。
        for kColumn in workbook_orgTable_all[objectiveSheet].iter_cols(min_col=iCell.column, 
                                                                         max_col=iCell.column,
                                                                         min_row=iCell.row+1, max_row=None):
            for kCell in kColumn:
                workbook_orgTable_all_dict[colNames[iCell.column-1]].append(kCell.value)
                      
        # iCell.column-1，报错TypeError: unsupported operand type(s) for -: 'str' and 'int'，我的2.4版本源代码中确实把column
        # 写成了获得单元格列字母的属性，col_idx是单元格列号的属性，到2.6将column改成了列号，获得列字母的属性改成了get_column_letter。
        # 这个报错得通过看代码来分析……
data = pd.DataFrame(workbook_orgTable_all_dict, columns=colNames)
originalData_jieruzongbiao=data.loc[:,:'zsjrsjc'].copy()
# columns参数应该是用不着的，dict中的键会被拿来当列名。
# openpyxl很聪明，把日期列的空单元格都转化成了NaT，把文本列的空单元格转化成了NAN。



# originalData_jieruzongbiao.where(originalData_jieruzongbiao==' ')
print(
    originalData_jieruzongbiao.loc[originalData_jieruzongbiao.isin([' ']).any(axis=1)].index,'\n',
    originalData_jieruzongbiao.loc[:,originalData_jieruzongbiao.isin([' ']).any(axis=0)].columns
)
# 侦测空格：这段代码打印df有空格的元素行列label，果然有空格。
originalData_jieruzongbiao.replace(to_replace=r'^[ \t\n\r\f\v]+|[ \t\n\r\f\v]+$',value=np.nan,regex=True,inplace=True)
# 如果to_replace, value参数中任何一个采取regex形式，需令regex=True。如果value=None，则不替换，注意这个，虽然None与np.nan
# 在df中涉及df运算时被同等对待，但在字符串操作比如替换中，就又不一样。或者直接认为None和nan在df中也是不一样的得了。

originalData_jieruzongbiao.fillna({'xysjc':pd.NaT,'pxsjc':pd.NaT,'jrcssjc':pd.NaT,'zsjrsjc':pd.NaT}
                                          ,inplace=True)
print(
    originalData_jieruzongbiao.loc[originalData_jieruzongbiao.isin([' ']).any(axis=1)].index,'\n',
    originalData_jieruzongbiao.loc[:,originalData_jieruzongbiao.isin([' ']).any(axis=0)].columns
)
originalData_jieruzongbiao.loc[:,['xysjc','pxsjc','jrcssjc','zsjrsjc']].dtypes 
# 验证时间戳类型。
# originalData_jieruzongbiao['zsjrsjc'].apply(lambda x: 0 if isinstance(x,(dt.datetime,pd._libs.tslibs.nattype.NaTType)) else 1)
# 验证是否都转化为了目标类型。apply对series直接执行elementwise操作（有sequence相关的函数对它做沿轴的操作），
# applymap对df执行elementwise操作。
# 后续如有其它脏数据，再改regex。



# 上月运行月报zsjrjg。
officiallyParticiOrg_jieruzongbiao=originalData_jieruzongbiao.loc[
    (originalData_jieruzongbiao['hzxy']!='4xyjc')
        &~(originalData_jieruzongbiao['zsjr'].isna())
           &(originalData_jieruzongbiao['zsjrsjc']<=dt.datetime(2021,8,31)),:
    ]
        
officiallyParticiOrg_jieruzongbiao_danweiquancheng=officiallyParticiOrg_jieruzongbiao.loc[:,'单位全称'
    ]
print(
    '截至{}月末机构{}家({:%Y%m%d})'.format
    (
        dt.date.today().month-2,
        officiallyParticiOrg_jieruzongbiao.shape[0],dt.datetime.today()
    )
)
officiallyParticiOrg_jieruzongbiao_20210928=officiallyParticiOrg_jieruzongbiao.copy()



print(set(officiallyParticiOrg_jieruzongbiao_20220112['单位全称'])-set(officiallyParticiOrg_jieruzongbiao_20210928['单位全称']),'\n',
set(officiallyParticiOrg_jieruzongbiao_20210928['单位全称'])-set(officiallyParticiOrg_jieruzongbiao_20220112['单位全称']))



# 上期运行月报对应的接口接入
originalData_jiekoujierubiao = pd.read_excel(
    r'D:\hlwjrxh\...\6-T+1jkbsjdb20210617.xlsx',
    sheet_name=0,header=0, names=None, index_col=None, 
    usecols=None, squeeze=False, dtype=None, engine=None, 
    converters=None, true_values=None, false_values=None, skiprows=None, nrows=None, 
    na_values=None, keep_default_na=True, verbose=False, 
    parse_dates=False, date_parser=None, thousands=None, comment=None, 
    skipfooter=0, convert_float=True, mangle_dupe_cols=True
)

# 1224后续要读取两个月份的jkbsjdb，发现还是把更新日期列统一成datetime64更方便。通过如下代码：
# originalData_jiekoujierubiao.loc[
#     originalData_jiekoujierubiao['更新日期'].str.contains(r'^[ \t\n\r\f\v]+|[ \t\n\r\f\v]+$', regex=True).notna(),:]
# 检测出某个单元格是string且不含whitespace，是False，非string的元素都返回NaN，这应该就是字符了，输出发现果然是备注列填到了更新日期列。
officiallyParticiOrg_jiekoujierubiao=originalData_jiekoujierubiao.loc[
    (originalData_jiekoujierubiao['进展情况']=='已完成')
    &(
        (originalData_jiekoujierubiao['更新日期'].isna())
        |(originalData_jiekoujierubiao['更新日期']<=dt.datetime(2021,8,31))
    ),:
]
# pd.NaT>dt.datetime(2020,11,30)，无论>, <,都返回False，不可比较，与nan类似。
# copy出来后面更新cxzb时用。
officiallyParticiOrg_jiekoujierubiao_20210617=officiallyParticiOrg_jiekoujierubiao.copy()
print(
    '截至{}月末机构{}家，机构{}家({:%Y%m%d})'.format
    (
        dt.date.today().month-2,
        officiallyParticiOrg_jiekoujierubiao.shape[0],
        officiallyParticiOrg_jieruzongbiao_20220112.shape[0]-officiallyParticiOrg_jiekoujierubiao.shape[0],
        dt.datetime.today()
    )
)
# originalData_jiekoujierubiao.dtypes

# print与左括号需要在一行里，否则下面再有statement或者expression，print的结果就打印不出来。如果print(，则print函数正常执行。
# 如果print
# ()，这样就会被解读为一个print builtin function，括号里的就是一个独立变量，print(这种打印出来不带引号，print与左括号不在一行，
# 输出的是带引号的，是变量内容，不是打印内容。




# jkbsjdb上期本期一致性。
print(set(officiallyParticiOrg_jiekoujierubiao_20211227['机构全称'])-set(officiallyParticiOrg_jiekoujierubiao_20210617['机构全称']),'\n',
set(officiallyParticiOrg_jiekoujierubiao_20210617['机构全称'])-set(officiallyParticiOrg_jiekoujierubiao_20211227['机构全称']))
# 0517云盘中的jkbsjdb仍没变，与4月份一样。



# 如果本期上期jrzb、jkbsjdb都是本期包含上期，则执行此代码。
if (set(officiallyParticiOrg_jieruzongbiao_20220112['单位全称'])>=set(officiallyParticiOrg_jieruzongbiao_20210928['单位全称']))\
and (set(officiallyParticiOrg_jiekoujierubiao_20211227['机构全称'])>=set(officiallyParticiOrg_jiekoujierubiao_20210617['机构全称'])):
    print(
        set(officiallyParticiOrg_jieruzongbiao_20220112['单位全称'])-\
        set(officiallyParticiOrg_jieruzongbiao_20210928['单位全称']),'\n',
        (set(officiallyParticiOrg_jiekoujierubiao_20211227['机构全称'])-set(officiallyParticiOrg_jiekoujierubiao_20210617['机构全称']))
    )
else:
    print('Unexpected elements:','\n',
         set(officiallyParticiOrg_jieruzongbiao_20210928['单位全称'])-\
         set(officiallyParticiOrg_jieruzongbiao_20220112['单位全称']),
         set(officiallyParticiOrg_jiekoujierubiao_20210617['机构全称'])-set(officiallyParticiOrg_jiekoujierubiao_20211227['机构全称']))
# 后续可以加一段打标签代码，将本期新增机构打上网页、接口标签，方便插入到cxzb里。
# orgToAdd={'网页':None,'接口':None}



import time, re, string, os, pdb
# import random
import xlrd, openpyxl,xlwings


workbook_queryTable_all = xlwings.Book(r'D:\jlwjrxh\...'+
        r'\5-cxjrzb202109141120-zsjrschjmd275+46.xlsx')
#              read_only=None, format=None, password=None, write_res_password=None, 
#              ignore_read_only_recommended=None, origin=None, delimiter=None, editable=None, notify=None, 
#              converter=None, add_to_mru=None, local=None, corrupt_load=None, impl=None
# 这些参数都是0.11之后逐步加入的……
# workbook_queryTable_all.selection.formula2
# AttributeError: 'Range' object has no attribute 'formula2', 
# 0.11文档里没有formula2，0.22文档中有，what's new里面也没有记载formula2什么时候加的。
# xlwings对Excel的读取是实时的，openpyxl、xlrf、read_excel读取最近一次保存的。
objectiveSheet = workbook_queryTable_all.sheets['接入机构0914']

workbook_zhoubaoshuju=openpyxl.load_workbook(r'D:\hlwjrxh...\2021年8月'
                                             +'\\202109141801gxptzbsj-仅更新各类接入机构名单275+46.xlsx',
                                            data_only=True)
# 输出最新的周报数据工作表名字，免得更新J列公式时手打。用openpyxl，用xlwings会把工作簿打开导致公式产生变化，
# 导致regex失效。
print(workbook_zhoubaoshuju.sheetnames)
# 20210317发现没写更新周报数据工作簿公式的代码。后续补充吧……



# 以下为获取数据区域右下角单元格坐标的方法。平时用solution2即可，solution1有些慢。

# # solution1

# # lastCellContainData参考https://stackoverflow.com/questions/33418119/xlwings-function-to-find-the-last-row-with-data
# # Stefan的答案。
# def lastCellContainData(objectiveSheet,lastRow=None,lastColumn=None):
#     lastRow = objectiveSheet.cells.last_cell.row if lastRow==None else lastRow
#     lastColumn = objectiveSheet.cells.last_cell.column if lastColumn==None else lastColumn
#     lastRows,lastColumns = [],[]
#     for col in range(1,lastColumn):
#         lastRows.append(objectiveSheet.range((lastRow, col)).end('up').row)
#         # 这里涉及排序算法，26^3列，17000列左右，是把每一列的最大行得出后进行max，还是每到下一列就进行一次比较，
#         # 把小于上一列最大行的行数直接赋值为1？这个也会涉及max的算法复杂度。当然搞一对lastRow，lastColumn变量会把这个问题简化。
    
#     for row in range(1,lastRow):
#         lastColumns.append(objectiveSheet.range((row, lastColumn)).end('left').column)
#         # 这里涉及排序算法，26^3列，17000列左右，是把每一列的最大行得出后进行max，还是每到下一列就进行一次比较，
#         # 把小于上一列最大行的行数直接赋值为1？这个也会涉及max的算法复杂度。当然搞一对lastRow，lastColumn变量会把这个问题简化。
#     return max(lastRows),max(lastColumns)
    
# lastCellContainData(objectiveSheet,lastRow=500,lastColumn=30)
# # xlwings这点非常不好啊，还得确定数据区域，而且如果源工作簿处于筛选状态下，也会读取全部数据，但显示的数据仅仅是筛选状态下的数据，
# # 即用end方法，使用范畴只能是被筛选出来的数据，果然是a smart wrapper of pywin32。

# solution2
objectiveSheet.range('A1:'+'AB500').current_region.rows.count,objectiveSheet.range('A1:AB1000').current_region.columns.count


# 取列号，列名，第一行的值，得到含有公式的列是哪几列，每列的外部工作簿是什么。
# 执行时注意，需要把外部工作簿关闭，否则cxzb公式中外部工作簿会变成相对引用。可能因为装在内存中了。

objectiveSheet.range((1,1),(1,18)).get_address(row_absolute=False, column_absolute=False)
columnHeaderContent = list(zip(
    string.ascii_uppercase,objectiveSheet.range((1,1),(1,18)).value,list(objectiveSheet.range((2,1),(2,18)).formula[0])
))
columnHeaderContentDict={}
for element in columnHeaderContent:
    columnHeaderContentDict[element[0]]=element[1:]
# 这个可以用来查看raw value，原始值是什么。方便debug。
# 搞成dict，有利于后面程序应用。

externalReferenceWb,path,keyword,currentWb = dict(),dict(),dict(),dict()
for column,tuples in columnHeaderContentDict.items():
    if tuples[-1].startswith('='):
        # search, match are different. 这里要找出所有匹配的还是得用findall。
#         print(re.findall(r'(\[.+?\])',tuples[-1]))
        # 经试验，如果regex中没有group，则调用groups返回()，调用group(1)返回IndexError。
        # for Wb in re.findall(r'(\[.+?\])',tuples[-1]):
        # 注意外部链接的工作簿要关掉，否则公式不显示绝对路径。
        path[column]=re.findall(r'(\'.+?\[)',tuples[-1])[0].strip('\'[')
        externalReferenceWb[column]=set([elements.strip('[].xlsx') for elements in re.findall(r'(\[.+?\])',tuples[-1])])
        
        rawSet=set(re.findall(r'[\u4e00-\u9fff]+|[A-Za-z]+',str(re.findall(r'(\[.+?\])',tuples[-1]))))
        # 会把xls，xlsx也找出来。不好在[A-Za-z]中去掉。
        rawSet.discard('xls')
        rawSet.discard('xlsx')
        keyword[column]=rawSet
        print('{0:<5}{1:#<{width}s}{2:}'.format(column,tuples[0]
                                                ,set([elements.strip('[].xlsx') for elements in re.findall(r'(\[.+?\])',tuples[-1])])
                                                ,width=25))
        # 经试验，格式化打印中，如果field_name，format_spec中的fill、align这几个变量组合，出现了汉字ascii混杂，
        # 就会出现对不齐，width是字符数，比如
        # '{0:<5}{1:#<{width}s}{2:s}'.format(tuples[0],tuples[1],'gggggggg顶顶顶顶顶',width=30)，tuples[1]是四个汉字，则
        # 剩下的空间由26个#填充（汉字ascii混杂），3个汉字则27个#填充，但屏幕显示时，#长度是汉字的一半，
        # 所以屏幕上看30个字符的长度就不一样。
        # list不能当string输出，{2:s}是不能在format()中对应list的，
        # 否则报错TypeError: unsupported format string passed to list.__format__
        currentWb[column]=list(set([elements.strip('[].xlsx') for elements in re.findall(r'(\[.+?\])',tuples[-1])]))

# 最终会打印externalReferenceWb，如果想打印path，keyword，继续调整print即可。



# 得到修改日期最新的业务表格。
# 此段代码对应folderKeyWordPairDict的数据结构为：值为list的dict。
# folderKeyWordPairDict没法按照列字母展开，因为dict的key不能重复，
# ['jrzb','查询jrzb','jkbsjdb']这三张表在同一个路径下，用dict只能对应1个path key，重复的key对应的value都被更新了。
# for dir_name,subdirs,files in os.walk(r'D:\hlwjrxh\...'):
#     print(dir_name,subdirs,files)
# 注意对汉字的排序是根据编码中的code number，不是拼音顺序。
folderKeyWordPairDict = {
    r'D:\jlwjrxh\...\cyzb':['1-jrzb','5-查询jrzb','6-T+1jkbsjdb'],
    r'D:\jlwjrxh\...\zbxzhybzz\2021年8月':['gxptzbsj'],
    r'D:\jlwjrxh\...\20210811':['dataPerMonth','successedUploadRatio','updateRatioTable','各月业务量']
}
def newestWorkbook(directory=r'.',workbookKeyWord=''):
    dirContent = os.listdir(directory)
    
    newestWorkbookName = []
    for keyWord in workbookKeyWord:
        sampleList = []
        for item in dirContent:
            if (keyWord in item) and (not item.startswith('~$')) and (item.endswith(('xlsx','xls'))):
                # 注意这里别用tilde表示not，tilde表示bitwise not是pandas里的规则。python里
                # bitwise negation operator: it takes the number n as binary number and “flips” 
                # all bits 0 to 1 and 1 to 0 to obtain the complement binary number.
                sampleList.append(item)
#                 print(sampleList)
# sampleList为空肯定是workbookKeyWord内容出了问题
#                 pdb.set_trace()
        sampleList.sort()
        try:
            newestWorkbookName.append(sampleList[-1])
            # 如果没有，sampleList是[]，取-1就会indexError，还是加个try。
        except IndexError:
            print('Probably folder {0} doesn\'t contain files {1} in Keyword list.'.format(directory,keyWord))
    return newestWorkbookName

for folder,KeyWord in folderKeyWordPairDict.items():
    print(newestWorkbook(
        directory=folder,workbookKeyWord=KeyWord)
         )
    


# 执行前要去掉筛选，否则被筛选掉的行，没法更新公式。
def changeFormula(column=None,rowRange=None,externalWbPath=None,currWb=None,newestWorkbookNamePaste=None,
                 pattern=None,replace=None):
    '''
    cnadidate variable of currWb: 
     currentWb=
     {'F': ['1-jrzb20210119'],
     'G': ['1-jrzb20210119'],
     'H': ['1-jrzb20210119'],
     'J': ['202012181800gxptzbsj-仅更新各类接入机构名单275+46'],
     'O': ['202009301055dataPerMonth'],
     'P': ['202009301100successedUploadRatio'],
     'Q': ['202009301057updateRatioTable'],
     'R': ['1-1各月业务量(2020-10-09)']}
     workbookKeyWord=currentWb['F'][0]
     
     newestWorkbookNamePaste:
         如果仅需替换公式中工作簿名，则这个参数是有用的。但如果路径和工作簿名都变了，这个参数没用。还是replace适用范围更广。
         you'd better paste the newest workbook name directly. Later I will improve newestWorkbook() to return a dict 
         like externalWbPath, so the pasting can be replaced by variable passing. newestWorkbookNamePaste here is of the same 
         function as replace. 
     replace:
         More generic than newestWorkbookNamePaste.
     '''
    # newestWorkbook传递进来也不好处理。怎么填充到pattern中？前面的序号1-、5-这些
    if newestWorkbook==None:
        print('No newestWorkbook as {workbookKeyWord}.'.format(workbookKeyWord=currWb))
        return
    # objectiveSheet.range((1,1),(459, 18)).options(pd.DataFrame,header=1).formula
    # 这种方式没法提取formula，只能得到值。
    # 看来是没办法用列名去提取公式了，只能用列序号。
    # https://stackoverflow.com/questions/48513093/xlwings-how-to-select-an-entire-column-without-headers
    # wb.sheets[0].range('A:A')[1:].value 这种会把last cell with data后面的空单元格全部返回，
    # Alternatively, define an Excel Table Object (Insert > Table):  wb.sheets[0].range('Table1[[#Data]]').value
    # 这个涉及表格对象，插入表格的那种表格，需要先构造表格对象，再使用。
#     newestWorkbook=newestWorkbook.strip('.xlsx')
    formulaTuple = objectiveSheet.range('{column}{rowRangeStart}:{column}{rowRangeEnd}'.format(
        column=column,rowRangeStart=rowRange[0],rowRangeEnd=rowRange[1])
                                       ).formula
#     pdb.set_trace()
#     pattern = re.compile(r'\[{workbookKeyWord}20[0-9]{{2}}[0-1][0-9][0-3][0-9]'.format(workbookKeyWord=workbookKeyWord))
#     pattern = columnHeaderContentDict['F'][-1].split('\'')[1]
    newFormula = [re.sub(pattern,replace, j, count=0, flags=0) for i in formulaTuple for j in i]
    # 不用在regex中加caret，1-jrzb并不在公式string的开头。注意{{}}，specifier的escape，参考Format String Syntax或者
    # https://stackoverflow.com/questions/1875676/python-2-6-str-format-and-regular-expressions
    # newFormula = [j.replace('[1-jrzb20201110.xlsx]','[1-jrzb20201230.xlsx]') for i in formulaTuple for j in i]
    # 这个不是很flexible，改代码或者输错jrzb文件名日期、更新jrzb文件名日期时比较麻烦。
    # 注意nested list comprehension，就是把内外循环放平：
    # for i in formulaTuple：
    #     for j in i：
    #         j
    # list应该可以直接赋值到单元格里……
    return newFormula

rowRange=(2,463)
newestWorkbookNamePaste='1-jrzb20210910.xlsx'
pattern=r'\[.+\..+\]'
replace = '['+newestWorkbookNamePaste+']'
for column in ['F','G','H']:
    
#     replace = columnHeaderContentDict[columns][-1].split('\'')[1]
    newFormula= changeFormula(
        column=column,
        rowRange=rowRange,
        externalWbPath=None,
        currWb=currentWb[column],
        newestWorkbookNamePaste=newestWorkbookNamePaste,
        pattern=pattern,replace=replace)
#     pdb.set_trace()
    print(newFormula[0])
    if '1-jrzb20210910.xlsx' in newFormula[0]:
        objectiveSheet.range('{column}{rowRangeStart}:{column}{rowRangeEnd}'.format(
            column=column,rowRangeStart=rowRange[0],rowRangeEnd=rowRange[1])
                                           ).formula = newFormula
# 20210318对于变量单复数，得定个开发规则，今天debug，公式赋值，三列都赋给一列了，经过set_trace()排错，发现每步都是正确在推进的。
# 最大的嫌疑落在了column相关的变量上。果然最后一个column变量用的是之前columns的值……肯定列号就不动了。原则就是遵从正常语法，
# 比如for element in some_list，就用单数，因为本身拿出来循环就是拿出来一个元素，如果是tuple这种保留字，for里面就用复数吧。




# for columns in ['J']:
#     newFormula= changeFormula(column=columns,rowRange=(2,459),externalWbPath=None,
#                   workbookKeyWord='gxptzbsj',newestWorkbook='202101251800gxptzbsj-仅更新各类接入机构名单275+46.xlsx')
#     if newFormula!=None:
#         objectiveSheet.range('{column}{rowRangeStart}:{column}{rowRangeEnd}'.format(
#             column=columns,rowRangeStart=rowRange[0],rowRangeEnd=rowRange[1])
#                                            ).formula = newFormula

# 注意执行前Excel要去除所有筛选，否则只更新被筛选出来的行……
column='J'
formulaTuple = objectiveSheet.range('{column}{rowRangeStart}:{column}{rowRangeEnd}'.format(
    column=column,rowRangeStart=rowRange[0],rowRangeEnd=rowRange[1])
                                   ).formula
# print(formulaTuple)
pattern=r'2021年8月\\\[202109141801gxptzbsj\-仅更新各类接入机构名单275\+46\.xlsx\]0715zzqjkjr'
# 转义序列运算优先级最高，所以3个backslash，第一个转义第二个，第二个转义下一个字符，最终意思是文本意义的\[。当左括号被正确转义后，
# 中间的连词符hyphen也就不用转义了，为了保险转义一下也行。pattern是正则，replace是文本。
replace=r'2021年8月\[202109141801gxptzbsj-仅更新各类接入机构名单275+46.xlsx]0715zzqjkjr'
newFormula = [re.sub(pattern,replace, j, count=0, flags=0) for i in formulaTuple for j in i]
print(newFormula[0])
if (newFormula!=None) and ('2021年8月' in newFormula[0]):
    objectiveSheet.range('{column}{rowRangeStart}:{column}{rowRangeEnd}'.format(
        column=column,rowRangeStart=rowRange[0],rowRangeEnd=rowRange[1])
                                       ).formula = newFormula
# formula赋值操作，报错只有一个com error，很多时候需要print(), pm(), set_trace()协同操作来debug，而且涉及公式批量操作，
# 有时候打开外部工作簿，Excel公式就会变化（绝对路径不见了），这时xlwings的读取也会产生变化，查找替换就会失效，导致公式
# 最终赋值不成功。对于J列更新，涉及路径变化、工作簿变化、工作表变化，处理时要小心。
# J列这种从路径到工作表都变化的情况，到时候读取一下gxptzbsj，输出一下对应的最新工作表，
# 直接复制最新的工作表名字，避免手打。
# 20210615试了一下这段代码可以兼容周报数据工作簿打开时的情形，因为替换操作中不含路径，所以不受工作簿打开状态影响。



for columns in ['O','P','Q','R']:
    newFormula= changeFormula(column=columns,rowRange=(2,459),externalWbPath=None,
                  workbookKeyWord=keyword[columns],newestWorkbook=None)
    if newFormula!=None:
        objectiveSheet.range('{column}{rowRangeStart}:{column}{rowRangeEnd}'.format(
            column=columns,rowRangeStart=rowRange[0],rowRangeEnd=rowRange[1])
                                           ).formula = newFormula
# 'O','P','Q','R'先不更新了，反正202010月改称申请考核后才逐家进行考核，不再统一考核了，这四列也就没有更新过。



# 还得写一下重命名工作表的代码。重命名一下工作表。
workbook_queryTable_all.save(
    r'D:\hlwjrxh\...\5-查询jrzb{:%Y%m%d%H%M}-zsjrschjmd275+46.xlsx'.format(
        dt.datetime.today()
    )
)



# 读取cxzb数据。openpyxl比xlwings快，而且不会打开被读取的工作簿。
# 既然用openpyxl打开，记得保存了更改的cxzb后再来读。
workbook_queryTable_all=openpyxl.load_workbook(r'D:\hlwjrxh\...'
                                               +r'\\5-查询jrzb202201121553-zsjrschjmd275+46.xlsx',
                                            data_only=True)
sheetNames = workbook_queryTable_all.sheetnames
print(sheetNames)
objectiveSheet = '接入机构0112'
nCols = workbook_queryTable_all[objectiveSheet].max_column
nRows = workbook_queryTable_all[objectiveSheet].max_row
print(nRows,nCols)

colNames = []
workbook_queryTable_all_dict = {}
for jRow in workbook_queryTable_all[objectiveSheet].iter_rows(min_row=None, max_row=nRows, min_col=None, max_col=nCols, 
                                                            ):
    # values_only=False这个参数2.6及以上才有，我目前是2.4。
    for jCell in jRow:
        if jCell.value:
            # library CH4 定义了哪些object会被视为False。
            # 我没用worksheet的values属性，因为values是按行返回值，
            # 但没有对偶的按列返回值的属性，如果遇到了按列返回值的需求还得重新写循环。
            print(jCell.row)
            # 注意openpyxl计数大部分是1-base。
            jCell_value=jCell.value
            break
    if jCell.value:
        break
    # 内层for中放2个break不管用，外层for的末尾放break会让外层for运行完第一次就提前break，得加个判断。
# 判定从哪行开始有数据，不能用workbook_orgTable_all[objectiveSheet].min_row，用各月业务量测试，这个返回1，显然第一行没数据。
# 不知道min_row对有无数据是怎么判定的。
for iRow in workbook_queryTable_all[objectiveSheet].iter_rows(min_col=None, max_col=None, min_row=jCell.row, 
                                                                         max_row=jCell.row):
    # 迭代出表头这一行，将这一行单元格的值append到colNames中。
    # iter_rows很灵活的方法。
    for iCell in iRow:
        colNames.append(iCell.value)
        # 先给字典每个键值赋一个空list，再append，[].append不返回任何值，不能直接赋值给其他变量。
        workbook_queryTable_all_dict[colNames[iCell.column-1]] = []
        # 对于表头每个cell所在的列，把这一列的值append到字典中。不能用list comprehension，否则把嵌套list赋给字典了。
        # 还是得套循环，这不如xlrd直接用col_values返回list方便。
        for kColumn in workbook_queryTable_all[objectiveSheet].iter_cols(min_col=iCell.column, 
                                                                         max_col=iCell.column,
                                                                         min_row=iCell.row+1, max_row=None):
            for kCell in kColumn:
                workbook_queryTable_all_dict[colNames[iCell.column-1]].append(kCell.value)
                      
        # iCell.column-1，报错TypeError: unsupported operand type(s) for -: 'str' and 'int'，我的2.4版本源代码中确实把column
        # 写成了获得单元格列字母的属性，col_idx是单元格列号的属性，到2.6将column改成了列号，获得列字母的属性改成了get_column_letter。
        # 这个报错得通过看代码来分析……
        # 20201229突然想到，之前搜到过Worksheet.cell()的用法，只提供row，column参数，就是读取某个单元格，再使用value属性就能
        # 获得单元格的值，所以这段读取工作簿的代码，也可以用Worksheet.cell(row,column).value的思路重写一下，看看是否比现在代码简单。
data = pd.DataFrame(workbook_queryTable_all_dict, columns=colNames)
originalData_queryTable=data.copy()
# columns参数应该是用不着的，dict中的键会被拿来当列名。
# openpyxl很聪明，把日期列的空单元格都转化成了NaT，把文本列的空单元格转化成了NAN。



# 侦测空格：

# originalData_jieruzongbiao.where(originalData_jieruzongbiao==' ')
print(
    originalData_queryTable.loc[originalData_queryTable.isin([' ']).any(axis=1),:].index,'\n',
    originalData_queryTable.loc[:,originalData_queryTable.isin([' ']).any(axis=0)].columns
)
# 这段代码打印df有空格的元素行列label，果然有空格。


# 没空格替换就不用执行。
originalData_queryTable.replace(to_replace=r'^[ \t\n\r\f\v]+|[ \t\n\r\f\v]+$',value=np.nan,regex=True,inplace=True)
# 如果to_replace, value参数中任何一个采取regex形式，需令regex=True。如果value=None，则不替换，注意这个，虽然None与np.nan
# 在df中涉及df运算时被同等对待，但在字符串操作比如替换中，就又不一样。或者直接认为None和nan在df中也是不一样的得了。

originalData_queryTable.fillna({'xysjc':pd.NaT,'pxsjc':pd.NaT,'jrcssjc':pd.NaT,'zsjrsjc':pd.NaT}
                                          ,inplace=True)
print(
    originalData_queryTable.loc[originalData_queryTable.isin([' ']).any(axis=1)].index,'\n',
    originalData_queryTable.loc[:,originalData_queryTable.isin([' ']).any(axis=0)].columns
)
originalData_queryTable.loc[:,['xysjc','pxsjc','jrcssjc','zsjrsjc']].dtypes 
# 验证时间戳类型。
# originalData_jieruzongbiao['zsjrsjc'].apply(lambda x: 0 if isinstance(x,(dt.datetime,pd._libs.tslibs.nattype.NaTType)) else 1)
# 验证是否都转化为了目标类型。apply对series直接执行elementwise操作（有sequence相关的函数对它做沿轴的操作），
# applymap对df执行elementwise操作。
# 后续如有其它脏数据，再改regex。


# 加个时间戳。
originalData_queryTable_20220112=originalData_queryTable.copy()


# zsjrjg。
officiallyParticiOrg_queryTable=originalData_queryTable_20220112.loc[
    (originalData_queryTable['接入状态']!='4xyjc'),:
    ]
# officiallyParticiOrg_jieruzongbiao_danweiquancheng=officiallyParticiOrg_jieruzongbiao.loc[:,'单位全称'
#     ]
print(
    '截至{}月末zsjr机构{}家({:%Y%m%d})'.format
    (
        dt.date.today().month-1,
        officiallyParticiOrg_queryTable.shape[0],dt.datetime.today()
    )
)
# 202102171658这张cxzb，是由20201126那个cxzb复制过来的，没有新增记录，所以zsjrjg应该会比12月月报要少几家，最终测试时
# 再验证数据。



originalData_jierujigouhuizongbiao = pd.read_excel(
    r'D:\hlwjrxh\...\202201272055gxptdqjrqkhz.xlsx',
    sheet_name=0,header=0, names=None, index_col=None, 
    usecols=None, squeeze=False, dtype=None, engine=None, 
    converters=None, true_values=None, false_values=None, skiprows=None, nrows=None, 
    na_values=None, keep_default_na=True, verbose=False, 
    parse_dates=False, date_parser=None, thousands=None, comment=None, 
    skipfooter=0, convert_float=True, mangle_dupe_cols=True
)
# zsjrjg（jrhzb）
officiallyParticiOrg_jierujigouhuizongbiao=originalData_jierujigouhuizongbiao.loc[
    originalData_jierujigouhuizongbiao['sfjcxy']!='4xyjc',:
]
officiallyParticiOrg_jierujigouhuizongbiao_jierujigouquancheng=officiallyParticiOrg_jierujigouhuizongbiao.loc[:,'接入机构全称']




queryOrg_chaxunzongbiao=officiallyParticiOrg_queryTable.loc[
    (officiallyParticiOrg_queryTable['ktzt']=='已开通')
    &((officiallyParticiOrg_queryTable['tkcxsj'].isna())
      |(officiallyParticiOrg_queryTable['tkcxsj']<=dt.datetime(2021,12,31)
       )
     ),:
]

queryOrg_jierujigouhuizongbiao=officiallyParticiOrg_jierujigouhuizongbiao.loc[
    officiallyParticiOrg_jierujigouhuizongbiao['sfktsccx']=='已开通',:]

print('截至{}月底机构{}家.\n其中{}家'.format(
    dt.date.today().month-1,
    queryOrg_jierujigouhuizongbiao.shape[0],
    queryOrg_jierujigouhuizongbiao.loc[
        (queryOrg_jierujigouhuizongbiao['shifouhuiyuan或是否guanlian']=='是')
        |(queryOrg_jierujigouhuizongbiao['shifouhuiyuan或是否guanlian']=='guanlian')
        ,:].shape[0])
     )



queryOrgDataframe={'':[queryOrg_chaxunzongbiao,queryOrg_jierujigouhuizongbiao]
                   }
orgTypeColomn={'jrzb':'是否持牌（0921',
               'jrhzb':'机构类型'
                                }
orgTypeAndMemberTypeValueColomn={'jrzb':
                                 ('是否持牌（0921','',['是']),
                                 'jrhzb':
                                 ('机构类型','',['是','guanlian'])
                                }

# memberTypeColumn={'jrzb':,'jrhzb':
#                  }
memberType=['是','','guanlian']
orgType=['cpxjjg','hlwxfjr','jys','wlxd','xedk','xt','zdjg','yh']



# 1210循环写法
for dataframe in queryOrgDataframe['']:
#     print(dataframe.head(1))
    # 对于每张机构接入表格。
    orgCount,memberCount,relatedMemberCount=[],[],[]
    for keys,values in orgTypeAndMemberTypeValueColomn.items():
#         print(values)
        # 对于每张表格中的机构类型列。
        # dataframe可能与values发生错配，需要try，忽略错配形成的KeyError。
        try:
            print(dataframe[values[0]].name,dataframe[values[1]].name)
            for types in orgType:
                # 对于任一表格中，与对应的机构类型列，任一业态的统计数据。
                orgCount.append(dataframe.loc[dataframe[values[0]]==types,:].shape[0])
                # 会员这里两张表中对应列的值不是一套，所以还得写很多判断、try catch，与其在代码中嵌套很多层，
                # 不如在数据字典中多做一些集约化，或者compact design。
                # 把orgTypeColomn扩充为orgTypeAndMemberTypeValueColomn是个非常好的创意，将代码复杂性转移到数据结构上。
                memberCount.append(
                    (dataframe.loc[
                        (dataframe[values[0]]==types)
                        &(dataframe[values[1]]==values[2][0]),:].shape[0])
                )
#                 print(keys,types,memberCount[-1])
                try:
                    relatedMemberCount.append(
                                        (dataframe.loc[
                            (dataframe[values[0]]==types)
                            &(dataframe[values[1]]==values[2][1]),:].shape[0])
                    )
                except IndexError:
                    pass
                print('{keys}:{types}:{num},{memberCountAsOfType},{relatedMemberCount}'.format(
                    keys=keys,types=types,num=orgCount[-1],
                    memberCountAsOfType=memberCount[-1],
                    relatedMemberCount=relatedMemberCount[-1] if len(relatedMemberCount)>0 else 'None',
                    width=30)
                     )
                # relatedMemberCount由于jrzb中没有这一项统计，可能生成[]，所以需要加一个ifelse判断，否则index out of range。
            
        except KeyError as e:
            continue
        print('{keys}:{newline}{Dict}'.format(keys=keys,newline='\n',Dict=dict(zip(orgType,orgCount))),
             '\n',
             '{keys}memberCount:{newline}{Dict}'.format(keys=keys,newline='\n',Dict=dict(zip(orgType,memberCount))),memberCount,
              '\n',
              '{keys}relatedMemberCount:{newline}{Dict}'.format(keys=keys,newline='\n',Dict=dict(zip(orgType,relatedMemberCount)))
             )
        print(orgCount,memberCount,relatedMemberCount)
        # 我曹，1223发现一个问题，jrhzb的zhudai会员数量是3，但对应的表里明明是4个，我用loc筛选一下，确实是4个zhudai会员机构。
        # 怀疑是变量传递有问题，但打印的都是memberCount，分业态打印时，打印出来的数量是4个，是对的，用字典形式打印出来，就是3
        # 个，很可能问题出在形成字典的过程中。也排除了把jrzb中的zhudai会员数量3传递到jrhzb这里的可能（即df与统计的数量
        # 错配的可能）。试着把所有有关的变量都打印了一下，发现memberCount由于一直在被append，从内层循环到最外层，没有被清空过，
        # 所以最后它的长度是业态个数*2，所以进行zip的时候会出现不匹配，进而数据丢失或错配。
        # 我在最后增加了print(orgCount,memberCount,relatedMemberCount)，打印出memberCount，证实了这个猜测。
        # 把orgCount,memberCount,relatedMemberCount=[],[],[]从循环外面拿到第一层循环里面，可以解决这个问题。
        

        
        
queryOrg_chaxunzongbiao_benyue = officiallyParticiOrg_queryTable.loc[
    (officiallyParticiOrg_queryTable['ktzt']=='已开通')
    &(
          (officiallyParticiOrg_queryTable['tkcxsj']>=dt.datetime(2021,12,1))
          &(officiallyParticiOrg_queryTable['tkcxsj']<=dt.datetime(2021,12,30))
     )
    ,:]
# NaT与datetime相比较，都返回False。
queryOrg_chaxunzongbiao_jiekou = officiallyParticiOrg_queryTable.loc[
    (officiallyParticiOrg_queryTable['ktzt']=='已开通')
    &((officiallyParticiOrg_queryTable['tkcxsj'].isna())
      |(officiallyParticiOrg_queryTable['tkcxsj']<=dt.datetime(2021,12,30)
       )
     )
    &(
        (officiallyParticiOrg_queryTable['']=='')
        |(officiallyParticiOrg_queryTable['']=='、')
    ),:
]
queryOrg_chaxunzongbiao_benyue.shape,queryOrg_chaxunzongbiao_jiekou.shape



# 本月报数机构。
workbook_dataPerMonth = xlrd.open_workbook(
    r'D:\hlwjrxh\...\20220127\1-1各月业务量(2022-01-27).xls')
sheetNames = workbook_dataPerMonth.sheet_names()
print(sheetNames)
objectiveSheet = sheetNames[0]
# workbook_dataPerMonthdata = pd.DataFrame(workbook_dataPerMonth) 不能直接转化成df
nRows, nCols = workbook_dataPerMonth.sheet_by_name(objectiveSheet).nrows, workbook_dataPerMonth.sheet_by_name(objectiveSheet).ncols
colNames = []
workbook_dataPerMonthDict = {}
for j in range(0,nRows):
    if any(workbook_dataPerMonth.sheet_by_name(objectiveSheet).row_values(j)):
        print(j)
        break
for i in range(0, nCols):
    colNames.append(workbook_dataPerMonth.sheet_by_name(objectiveSheet).col_values(i)[j])
    workbook_dataPerMonthDict[workbook_dataPerMonth.sheet_by_name(objectiveSheet).col_values(i)[j]] = \
        workbook_dataPerMonth.sheet_by_name(objectiveSheet).col_values(i)[j+1:]
data = pd.DataFrame(workbook_dataPerMonthDict, columns=colNames)
dataPerMonth = data.copy()


dataPerMonth.loc[dataPerMonth['2021-11']>0,:].shape



import win32com.client

word = win32com.client.Dispatch("Word.Application")
word.visible = True
wb = word.Documents.Open(r'D:\hlwjrxh\...\20211109\B29998(2021-12-13).doc')
doc = word.ActiveDocument
text = doc.Range().Text
print(doc.Range().Text)
# https://stackoverflow.com/questions/36001482/read-doc-file-with-python
# 10SecTom的回答。直接复制，一执行就有了。
# print和直接查看Text，结果不一样。原因有空研究一下。
# 执行后，word老是占用资源，有一个不可见的word instance需要强制结束进程才会关闭。可能需要及时关闭相应工作簿。
# 20210517执行wb.Close后，关闭所有word工作簿，任务管理器中没有word了。quit报错。



wb.Close(SaveChanges=False)



monthlyOperationReportData = xlwings.Book(r'D:\hlwjrxh\....\2021年11月'+
        r'\202112212100运行月报数据.xlsx')
monthlyOperationReportData = xlwings.Book(r'D:\hlwjrxh\...\2021年11月'+
        r'\202112212100运行月报数据.xlsx')
# 想来想去，还是直接粘贴比较方便，搞字典或者df再操作来操作去，更麻烦。
objectiveSheet = monthlyOperationReportData.sheets['项目和人的粒度']

# current_region返回的是绝对引用，需提取最后一个被匹配的$+数字。last_cell也行，更简单。
baseRow=int(re.findall(r'\$[0-9]+',objectiveSheet.range('A5:'+'AB500').current_region.address)[-1].strip('$'))
# 之所以从第5行开始，因为第一此周报数据无法计算增量，增量那行空着，那里2*8的矩形会被认为是一个区域。



wichWeek = '1130周'
firstLine=[wichWeek,'自然人借款客户w','借款账户累计w','入库记录','余额y','网贷余额y','其他余额y','未结清借款账户数w']
# 如果需要入库记录调整字段，手动在Excel中加。
if firstLine[0]=='':
    raise Exception('Fill out the date in firstLine.')
number = re.findall(r'[0-9]+\.[0-9]{2,}',text)
number[0:0] = ['']
# 这是文档中对list进行insert的等价方式，还有list_a[:3]+elementToInsert+list_a[3:]这种方法。




# 复制到Excel中。
objectiveSheet.range('A{}'.format(
    baseRow+1
)
                    ).value=[firstLine,number]
objectiveSheet.range('B{}'.format(
    baseRow+3
)
                    ).value=['=b{}-b{}'.format(baseRow+2,baseRow-1),
                            '=c{}-c{}'.format(baseRow+2,baseRow-1),
                            '=d{}-d{}'.format(baseRow+2,baseRow-1)]




# 读取业务种类信息，由于这张表同时存在左上方的空行、空列，要加一段识别空列的代码。这个cell中的读取代码，
# 在处理multiIndex时，xlrd.load_workbook作用仅仅是获取空行空列行列号。
currentMonth_trafficByCategory_path = r'D:\hlwjrxh\...\2021年12月'+\
                                         r'\业务种类信息（全部和有效）(2022-01-29).xls'
lastMonth_trafficByCategory_path = r'D:\hlwjrxh\...\2021年11月'+\
                                         r'\业务种类信息（全部和有效）(2021-12-13).xls'
startRowColumnDict = {currentMonth_trafficByCategory_path:[],lastMonth_trafficByCategory_path:[]}
for path in [currentMonth_trafficByCategory_path,lastMonth_trafficByCategory_path]:
    trafficByCategory = xlrd.open_workbook(path,
                                             formatting_info=True)
    sheetNames = trafficByCategory.sheet_names()
    print(sheetNames)
    objectiveSheet = sheetNames[0]
    nRows, nCols = trafficByCategory.sheet_by_name(objectiveSheet).nrows, trafficByCategory.sheet_by_name(objectiveSheet).ncols
    # nrows属性，与xlwings中的current_region不一样，nrows应该是直接取了最右下的单元格计算的，不是取的有数据的最大矩形，而是
    # 把左上方的完全空白的列也囊括在内了。
    colNames = []
    colNames_Row1,colNames_Row2 = [],[]
    trafficByCategoryDict = {}
    for j_row in range(0,nRows):
        if any(trafficByCategory.sheet_by_name(objectiveSheet).row_values(j_row)):
            print(j_row)
            break
    startRowColumnDict[path].append(j_row)
    for j_col in range(0,nCols):
        if any(trafficByCategory.sheet_by_name(objectiveSheet).col_values(j_col)):
            print(j_col)
            break
    startRowColumnDict[path].append(j_col)
    for i_col in range(j_col, nCols):
        colNames.append(trafficByCategory.sheet_by_name(objectiveSheet).col_values(i_col)[j_row])
        trafficByCategoryDict[trafficByCategory.sheet_by_name(objectiveSheet).col_values(i_col)[j_row]] = \
            trafficByCategory.sheet_by_name(objectiveSheet).col_values(i_col)[j_row+1:]
            # 0217但这样不行，因为第一行的合并单元格中，有''作为列名，会重复，导致引用键值为''的键时，python会返回所有键值为''的键，
            # 导致赋值的时候这些键统一赋了值，比如先统一赋了有效D的值，下一个i就会赋有效I，以此类推，for循环走完的时候，所有键值为''的value
            # 都被赋予了最后一个月份的有效G的值。导致dict中键值数小于colNames长度，所以出现很多列都是有效+G的情况。
            # 还是得用read_excel去把业务数据这一部分作为一个multiIndex单独读取出来，再concat基本信息。

        # 0206这部分是为了下面生成multiIndex的几个level。
        # 没有混合level（某些列是2个level，某些列1个level）的multiIndex dataframe，没发现某些列1个level，某些列2个level这种创建方法，
        # 所以只能把前面序号、全称等列都用作第二个level，因为这几列对于每个月份是必有的。
        # 然后写代码拆解这些levels。现在打算弄成2行，如果某列对应的某行有空的，比如全称的下一行是空，则此列就自动算到第二个level
        # （第一个level的每列都有的这个level）中。如果一个level的某个label在每个index中都能找到，则肯定不是第一个level，否则就是第一个level
        # 比如序号，在每个index中都是有的，但2019年10月，只在部分index中能找到即（2019年10月，有效笔数）（2019年10月，有效未结清）等index中
        # 能找到，则是第一个level。——这个原则不对。应该改为如果某个level中所有的值，对于另一个level所有的值都出现了，则是下一级level。
        # ——也不对，笛卡尔积肯定两两level满足这个原则，这也是level互换级别的基础。所以月份和本月还款状态把哪个放到更高一级的level中都可以。
        # 原则应该是先把互为笛卡尔积的level找出来，月份与有效类型是，这两者与序号等几列也可以互为笛卡尔积。
        colNames_Row1.append(trafficByCategory.sheet_by_name(objectiveSheet).col_values(i_col)[j_row])
        colNames_Row2.append(trafficByCategory.sheet_by_name(objectiveSheet).col_values(i_col)[j_row+1])
        # 以后可以考虑把j+1一般化为j+k，容纳表头为3行及以上单元格合并而成的情况。

    data = pd.DataFrame(trafficByCategoryDict, columns=colNames)
    # 0217这样用dict生成dataframe不行，因为第一行的合并单元格中，有''作为列名，会重复，重复的列在df中会保留，但重复的键值在dict中会被覆盖，所以
    # 导致dict中键值数小于colNames长度，再生成df的时候会有一些列没有值，默认从上列复制过来，所以出现很多列都是有效+G的情况。
    # 还是得用read_excel去把业务数据这一部分作为一个multiIndex单独读取出来，再concat基本信息。
    # 20210303生成'业务种类信息（全部和有效）也会出现''作为列名的问题。

    # 0209怎么将excel中间的一堆数提取出来的，container选哪个？如果选df，columns怎样才能不重复？因为重复了在循环中就没法遍历，比如将excel
    # 内容直接读进df，很多列名=''，没法遍历来赋值。怎样通过循环遍历去赋值？可以先生成一个multiIndex df，然后用行的循环去将data中的值赋给
    # MultiIndex df，data中由于存在相同名字的列，所以只能按从左到右的顺序去赋值，可行，但不是那么stylish。

    # updateTableIndexNeeded = updateTableIndex.set_codes()
    # updateTable = pd.DataFrame(np.random.randn(data.shape[0]-1,len(cleanLevel1)*len(cleanLevel2)),columns=updateTableIndex)
    # 生成MultiIndex df的时候，read_excel虽然也能生成MultiIndex df，但似乎不适用于存在standAloneLabel的情况，见read_excel参数header的解释。
    # 0216可以通过codes参数来删掉不需要的multiIndex组合。先把basicInfo+cleanLevel1,standAloneLabel+cleanLevel2的笛卡尔积搞出来，再通过
    # list、series、tuple等对象的相关技巧，整理出相应的codes组合。
    # 后来发现用read_excel生成multiIndex，read_excel会自动填充standAloneLabel的下一级level，unnamed level，生成正常的df，
    # 如果需要有效项目数据，用slicing
    # 把它们抽出来就行了。不用把机构基本信息和有效项目数据拆开整理再concat。
print(startRowColumnDict)




# currentMonth_trafficByCategory_path
# 由于上月、当月业务种类信息起始行列可能不同，所以要把它们的起始行列都用startRowColumnDict记录下来，以备后面drop、set_index使用。
trafficByCategoryFromReadExcel_currPeriod = pd.read_excel(currentMonth_trafficByCategory_path,
                sheet_name=0,header=[startRowColumnDict[currentMonth_trafficByCategory_path][0],
                                     startRowColumnDict[currentMonth_trafficByCategory_path][0]+1], names=None, index_col=None, 
                usecols=None, squeeze=False, dtype=None, engine=None, 
                converters=None, true_values=None, false_values=None, skiprows=None, nrows=None, 
                na_values=None, keep_default_na=True, verbose=False, 
                parse_dates=False, date_parser=None, thousands=None, comment=None, 
                skipfooter=0, convert_float=True, mangle_dupe_cols=True)
trafficByCategoryFromReadExcel_lastPeriod = pd.read_excel(lastMonth_trafficByCategory_path,
                sheet_name=0,header=[startRowColumnDict[lastMonth_trafficByCategory_path][0],
                                     startRowColumnDict[lastMonth_trafficByCategory_path][0]+1], names=None, index_col=None, 
                usecols=None, squeeze=False, dtype=None, engine=None, 
                converters=None, true_values=None, false_values=None, skiprows=None, nrows=None, 
                na_values=None, keep_default_na=True, verbose=False, 
                parse_dates=False, date_parser=None, thousands=None, comment=None, 
                skipfooter=0, convert_float=True, mangle_dupe_cols=True)
# 如果在use_col中限定了具体的哪些列，则提示cannot specify usecols when specifying a multi-index header。可以先全读了，再slice。
# https://github.com/pandas-dev/pandas/issues/25449，解释了这个问题为什么一直没人处理，人力不够。
# 直到202103也没人处理。只能读取后丢弃了。把不区分有效无效的统计列drop掉。
# pdb.set_trace()
trafficByCategoryFromReadExcel_currPeriod.drop(
    columns=trafficByCategoryFromReadExcel_currPeriod.columns[0:startRowColumnDict[currentMonth_trafficByCategory_path][1]],
    inplace=True)
trafficByCategoryFromReadExcel_currPeriod.drop(columns='全部',level=1,inplace=True)
# 把业务种类列设为index，不然相减报错str不能相减。
trafficByCategoryFromReadExcel_currPeriod.set_index(keys=trafficByCategoryFromReadExcel_currPeriod.columns[0],inplace=True)

trafficByCategoryFromReadExcel_lastPeriod.drop(
    columns=trafficByCategoryFromReadExcel_lastPeriod.columns[0:startRowColumnDict[lastMonth_trafficByCategory_path][1]],
    inplace=True)
trafficByCategoryFromReadExcel_lastPeriod.drop(columns='全部',level=1,inplace=True)
trafficByCategoryFromReadExcel_lastPeriod.set_index(keys=trafficByCategoryFromReadExcel_lastPeriod.columns[0],inplace=True)
# 小数点位数太多，搞个条件格式监测数据变化。
growthRate_traffic=(trafficByCategoryFromReadExcel_currPeriod
            -trafficByCategoryFromReadExcel_lastPeriod)/trafficByCategoryFromReadExcel_lastPeriod
# 条件格式，见pandas documentation CH23 Styling章节。
# 代码从网上复制的，网上例子易懂。
def color_elements(val):
    if abs(val) <= 0.1:
        color = 'black'
    elif abs(val)<=0.2:
        color = 'blue'
    else:
        color ='red'
    return 'color:{}'.format(color)
growthRate_traffic.style.applymap(color_elements)




# 业务种类信息的总量数据需要粘贴。
objectiveSheet = monthlyOperationReportData.sheets['多头表的数据']
objectiveSheet.range('K4').options(transpose=True).value=list(
    trafficByCategoryFromReadExcel_currPeriod[('借款总人数','valityflag=\'0\'')])[0:-1]
objectiveSheet.range('M4').options(transpose=True).value=list(
    trafficByCategoryFromReadExcel_currPeriod[('借款总账户数','valityflag=\'0\'')])[0:-1]




# 读取业务种类信息，由于这张表同时存在左上方的空行、空列，要加一段识别空列的代码。这个cell中的读取代码，
# 在处理multiIndex时，xlrd.load_workbook作用仅仅是获取空行空列行列号。
currMonth_multiBorrTrafficByCategory_path = r'D:\hlwjrxh\...\2021年12月'+\
                                         r'\4-3(2022-01-29)12.xls'
lastMonth_multiBorrTrafficByCategory_path = r'D:\hlwjrxh\...\2021年12月'+\
                                         r'\4-3(2022-01-29)11.xls'
startRowColumnDict = {currMonth_multiBorrTrafficByCategory_path:[],lastMonth_multiBorrTrafficByCategory_path:[]}
for path in [currMonth_multiBorrTrafficByCategory_path,lastMonth_multiBorrTrafficByCategory_path]:
    multiBorrTrafficByCategory_xlrd = xlrd.open_workbook(path,
                                             formatting_info=True)
    sheetNames = multiBorrTrafficByCategory_xlrd.sheet_names()
    print(sheetNames)
    objectiveSheet = sheetNames[0]
    nRows, nCols = multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).nrows, \
    multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).ncols
    # nrows属性，与xlwings中的current_region不一样，nrows应该是直接取了最右下的单元格计算的，不是取的有数据的最大矩形，而是
    # 把左上方的完全空白的列也囊括在内了。
    colNames = []
    colNames_Row1,colNames_Row2 = [],[]
    multiBorrTrafficByCategory_xlrd_Dict = {}
    for j_row in range(0,nRows):
        if any(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).row_values(j_row)):
            print(j_row)
            break
    startRowColumnDict[path].append(j_row)
    for j_col in range(0,nCols):
        if any(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(j_col)):
            print(j_col)
            break
    startRowColumnDict[path].append(j_col)
    for i_col in range(j_col, nCols):
        colNames.append(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row])
        multiBorrTrafficByCategory_xlrd_Dict[multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row]] = \
            multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row+1:]
            # 0217但这样不行，因为第一行的合并单元格中，有''作为列名，会重复，导致引用键值为''的键时，python会返回所有键值为''的键，
            # 导致赋值的时候这些键统一赋了值，比如先统一赋了有效D的值，下一个i就会赋有效I，以此类推，for循环走完的时候，
            # 所有键值为''的value都被赋予了最后一个月份的有效G的值。
            # 导致dict中键值数小于colNames长度，所以出现很多列都是有效+G的情况。
            # 还是得用read_excel去把业务数据这一部分作为一个multiIndex单独读取出来，再concat基本信息。

        # 0206这部分是为了下面生成multiIndex的几个level。
        # 没有混合level（某些列是2个level，某些列1个level）的multiIndex dataframe，没发现某些列1个level，某些列2个level这种创建方法，
        # 所以只能把前面序号、全称等列都用作第二个level，因为这几列对于每个月份是必有的。
        # 然后写代码拆解这些levels。现在打算弄成2行，如果某列对应的某行有空的，比如全称的下一行是空，则此列就自动算到第二个level
        # （第一个level的每列都有的这个level）中。如果一个level的某个label在每个index中都能找到，则肯定不是第一个level，否则就是第一个level
        # 比如序号，在每个index中都是有的，但2019年10月，只在部分index中能找到即（2019年10月，有效笔数）（2019年10月，有效未结清）等index中
        # 能找到，则是第一个level。——这个原则不对。应该改为如果某个level中所有的值，对于另一个level所有的值都出现了，则是下一级level。
        # ——也不对，笛卡尔积肯定两两level满足这个原则，这也是level互换级别的基础。所以月份和本月还款状态把哪个放到更高一级的level中都可以。
        # 原则应该是先把互为笛卡尔积的level找出来，月份与有效类型是，这两者与序号等几列也可以互为笛卡尔积。
        colNames_Row1.append(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row])
        colNames_Row2.append(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row+1])
        # 以后可以考虑把j+1一般化为j+k，容纳表头为3行及以上单元格合并而成的情况。

    data = pd.DataFrame(trafficByCategoryDict, columns=colNames)
    # 0217这样用dict生成dataframe不行，因为第一行的合并单元格中，有''作为列名，会重复，重复的列在df中会保留，但重复的键值在dict中会被覆盖，所以
    # 导致dict中键值数小于colNames长度，再生成df的时候会有一些列没有值，默认从上列复制过来，所以出现很多列都是有效+G的情况。
    # 还是得用read_excel去把业务数据这一部分作为一个multiIndex单独读取出来，再concat基本信息。
    # 20210303生成'业务种类信息（全部和有效）也会出现''作为列名的问题。

    # 0209怎么将excel中间的一堆数提取出来的，container选哪个？如果选df，columns怎样才能不重复？因为重复了在循环中就没法遍历，比如将excel
    # 内容直接读进df，很多列名=''，没法遍历来赋值。怎样通过循环遍历去赋值？可以先生成一个multiIndex df，然后用行的循环去将data中的值赋给
    # MultiIndex df，data中由于存在相同名字的列，所以只能按从左到右的顺序去赋值，可行，但不是那么stylish。

    # updateTableIndexNeeded = updateTableIndex.set_codes()
    # updateTable = pd.DataFrame(np.random.randn(data.shape[0]-1,len(cleanLevel1)*len(cleanLevel2)),columns=updateTableIndex)
    # 生成MultiIndex df的时候，read_excel虽然也能生成MultiIndex df，但似乎不适用于存在standAloneLabel的情况，见read_excel参数header的解释。
    # 0216可以通过codes参数来删掉不需要的multiIndex组合。先把basicInfo+cleanLevel1,standAloneLabel+cleanLevel2的笛卡尔积搞出来，再通过
    # list、series、tuple等对象的相关技巧，整理出相应的codes组合。
    # 后来发现用read_excel生成multiIndex，read_excel会自动填充standAloneLabel的下一级level，unnamed level，生成正常的df，
    # 如果需要有效项目数据，用slicing
    # 把它们抽出来就行了。不用把机构基本信息和有效项目数据拆开整理再concat。
print(startRowColumnDict)




multiBorrTrafficByCategory_currPeriod = pd.read_excel(currMonth_multiBorrTrafficByCategory_path,
                sheet_name=0,header=[startRowColumnDict[currMonth_multiBorrTrafficByCategory_path][0],
                                     startRowColumnDict[currMonth_multiBorrTrafficByCategory_path][0]+1], 
                names=None, index_col=[startRowColumnDict[currMonth_multiBorrTrafficByCategory_path][1],
                                     startRowColumnDict[currMonth_multiBorrTrafficByCategory_path][1]+1], 
                usecols=None, squeeze=False, dtype={'金额逾期率':np.float64,'项目逾期率':np.float64}, engine=None, 
                converters={'金额逾期率':float,'项目逾期率':float}, true_values=None, false_values=None, skiprows=None, nrows=None, 
                na_values=None, keep_default_na=True, verbose=False, 
                parse_dates=False, date_parser=None, thousands=None, comment=None, 
                skipfooter=1, convert_float=True, mangle_dupe_cols=True)
multiBorrTrafficByCategory_lastPeriod = pd.read_excel(lastMonth_multiBorrTrafficByCategory_path,
                sheet_name=0,header=[startRowColumnDict[lastMonth_multiBorrTrafficByCategory_path][0],
                                     startRowColumnDict[lastMonth_multiBorrTrafficByCategory_path][0]+1], 
                names=None, index_col=[startRowColumnDict[lastMonth_multiBorrTrafficByCategory_path][1],
                                     startRowColumnDict[lastMonth_multiBorrTrafficByCategory_path][1]+1], 
                usecols=None, squeeze=False, dtype={'金额逾期率':np.float64,'项目逾期率':np.float64}, engine=None, 
                converters={'金额逾期率':float,'项目逾期率':float}, true_values=None, false_values=None, skiprows=None, nrows=None, 
                na_values=None, keep_default_na=True, verbose=False, 
                parse_dates=False, date_parser=None, thousands=None, comment=None, 
                skipfooter=1, convert_float=True, mangle_dupe_cols=True)
# multiBorrTrafficByCategory_currPeriod = pd.read_excel(r'D:\hlwjrxh\...\2021年4月'+
#                                          r'\4-3(2021-05-18)4月.xls',
#                 sheet_name=0,header=[j_row,j_row+1], names=None, index_col=[j_col,j_col+1], 
#                 usecols=None, squeeze=False, dtype={'金额逾期率':np.float64,'项目逾期率':np.float64}, engine=None, 
#                 converters={'金额逾期率':float,'项目逾期率':float}, true_values=None, false_values=None, skiprows=None, nrows=None, 
#                 na_values=None, keep_default_na=True, verbose=False, 
#                 parse_dates=False, date_parser=None, thousands=None, comment=None, 
#                 skipfooter=1, convert_float=True, mangle_dupe_cols=True)
# # skiprows是Rows to skip at the beginning (0-indexed).跳过开头几行。skipfooter是跳过最后的。
# multiBorrTrafficByCategory_lastPeriod = pd.read_excel(r'D:\hlwjrxh\...\2021年4月'+
#                                          r'\4-3(2021-05-18)3月.xls',
#                 sheet_name=0,header=[j_row,j_row+1], names=None, index_col=[j_col+1,j_col+2], 
#                 usecols=None, squeeze=False, dtype={'金额逾期率':np.float64,'项目逾期率':np.float64}, engine=None, 
#                 converters={'金额逾期率':float,'项目逾期率':float}, true_values=None, false_values=None, skiprows=None, nrows=None, 
#                 na_values=None, keep_default_na=True, verbose=False, 
#                 parse_dates=False, date_parser=None, thousands=None, comment=None, 
#                 skipfooter=1, convert_float=True, mangle_dupe_cols=True)
# 如果在use_col中限定了具体的哪些列，则提示cannot specify usecols when specifying a multi-index header。可以先全读了，再slice。
# https://github.com/pandas-dev/pandas/issues/25449，解释了这个问题为什么一直没人处理，人力不够。
# 直到202103也没人处理。只能读取后丢弃了。把不区分有效无效的统计列drop掉。

multiBorrTrafficByCategory_currPeriod.drop(columns='比上月增减',level=1,inplace=True)
# multiBorrTrafficByCategory_currPeriod.set_index(keys=multiBorrTrafficByCategory_currPeriod.columns[0:2],inplace=True)
# 老是提示TypeError: unsupported operand type(s) for -: 'str' and 'str'，逾期率有问题，用dtypes、converters都不行，原因在于
# 百分号，需要把它strip掉。用str属性不行，str只能应用于series和index，multiIndex loc取出来的2列形成df，不适用。
# apply(str.strip)不适用，传递过来的变量是series，没有strip方法，applymap只能应用接受一个参数的函数，且这个参数必须是df elements，
# 与strip的参数类型不符。只能替换了。
# 20210305这里不应该用np.nan，应该用''，因为to_replace指代的是符合regex的string，将这个string替换为value，
# 替换了也是object，只能astype({'金额逾期率':float,'项目逾期率':float})不管用，to_numeric就可以，我记得好几次了都是这样的。
# 如果取有特殊性质（如数据类型不同）的列，可以lambda。
multiBorrTrafficByCategory_currPeriod=multiBorrTrafficByCategory_currPeriod.apply(
    lambda x: pd.to_numeric(x.str.strip('%')) if x.dtype == "object" else x)
multiBorrTrafficByCategory_currPeriod.index.names=['category','NPL']

multiBorrTrafficByCategory_lastPeriod.drop(columns='比上月增减',level=1,inplace=True)
# multiBorrTrafficByCategory_lastPeriod.set_index(keys=multiBorrTrafficByCategory_lastPeriod.columns[0:2],inplace=True)
multiBorrTrafficByCategory_lastPeriod=multiBorrTrafficByCategory_lastPeriod.apply(
    lambda x: pd.to_numeric(x.str.strip('%')) if x.dtype == "object" else x)
multiBorrTrafficByCategory_lastPeriod.index.names=['category','NPL']

# 小数点位数太多，搞个条件格式监测数据变化。
growthRate_4_3=(multiBorrTrafficByCategory_currPeriod
            -multiBorrTrafficByCategory_lastPeriod)/multiBorrTrafficByCategory_lastPeriod
# 条件格式，见pandas documentation CH23 Styling章节。
# 代码从网上复制的，网上例子易懂。
def color_elements(val):
    if abs(val) <= 0.1:
        color = 'black'
    elif abs(val)<=0.2:
        color = 'blue'
    else:
        color ='red'
    return 'color:{}'.format(color)
growthRate_4_3.style.applymap(color_elements)




# 生成用4-3更新过后的业务种类信息，监测一下各业态之和的变化。
# 如何根据index level0 去汇总。sum？aggregate、apply，不用groupby，本来就grouped by 了。
# index.levels可以获得level0的unique value，doc没讲。
# category_Index = multiBorrTrafficByCategory_currPeriod.index.unique(level=0)
# 涉及借款人数的话，只要有分业态这个条件的数据都不能直接加总，会重复计算。分业态的借款账户数可以直接加总。
# 分平台个数统计的借款人数，直接加总的话，不会重复计算，原因在sql中，4-3sql，都是先算unique表，
# 先以记录的粒度把每条记录对应的借款人在几个平台借款统计出来，再分业态、平台个数来统计，所以同一个借款人有可能分散在不同的业态中。
# 搞了半天原来sum已经考虑了multiIndex的事情:
multiBorrTrafficByCategory_currPeriod_sum=multiBorrTrafficByCategory_currPeriod.sum(level='category')
multiBorrTrafficByCategory_currPeriod_sum.loc['总计',:]=multiBorrTrafficByCategory_currPeriod_sum.sum(axis=0)
# trafficByCategoryFromReadExcel_currPeriod_=trafficByCategoryFromReadExcel_currPeriod.copy()
trafficByCategoryFromReadExcel_currPeriod.index=multiBorrTrafficByCategory_currPeriod_sum.index
trafficByCategoryFromReadExcel_currPeriod[('未结清借款人数','valityflag=\'0\'')]=multiBorrTrafficByCategory_currPeriod_sum[
    ('借款人数','截至月末数值')]
trafficByCategoryFromReadExcel_currPeriod[('未结清账户数','valityflag=\'0\'')]=multiBorrTrafficByCategory_currPeriod_sum[
    ('借款笔数','截至月末数值')]
trafficByCategoryFromReadExcel_currPeriod[('余额','valityflag=\'0\'')]=multiBorrTrafficByCategory_currPeriod_sum[
    ('余额','截至月末数值')]
# 老是有NA为什么？当然partial indexing的时候要把所有level都写清。NA与partial indexing无关，用loc也是有NA，按理说只有总计那行有NA才对。
# 极有可能是跟index value不同有关，需要换一下index，使4-3与业务种类信息的index value保持一致。
multiBorrTrafficByCategory_lastPeriod_sum=multiBorrTrafficByCategory_lastPeriod.sum(level='category')
multiBorrTrafficByCategory_lastPeriod_sum.loc['总计',:]=multiBorrTrafficByCategory_lastPeriod_sum.sum(axis=0)
# trafficByCategoryFromReadExcel_currPeriod_=trafficByCategoryFromReadExcel_currPeriod.copy()
trafficByCategoryFromReadExcel_lastPeriod.index=multiBorrTrafficByCategory_lastPeriod_sum.index
trafficByCategoryFromReadExcel_lastPeriod[('未结清借款人数','valityflag=\'0\'')]=multiBorrTrafficByCategory_lastPeriod_sum[
    ('借款人数','截至月末数值')]
trafficByCategoryFromReadExcel_lastPeriod[('未结清账户数','valityflag=\'0\'')]=multiBorrTrafficByCategory_lastPeriod_sum[
    ('借款笔数','截至月末数值')]
trafficByCategoryFromReadExcel_lastPeriod[('余额','valityflag=\'0\'')]=multiBorrTrafficByCategory_lastPeriod_sum[
    ('余额','截至月末数值')]

growthRate_trafficOnLoanIncluded=(trafficByCategoryFromReadExcel_currPeriod
            -trafficByCategoryFromReadExcel_lastPeriod)/trafficByCategoryFromReadExcel_lastPeriod
# 条件格式，见pandas documentation CH23 Styling章节。
def color_elements(val):
    if abs(val) <= 0.1:
        color = 'black'
    elif abs(val)<=0.2:
        color = 'blue'
    else:
        color ='red'
    return 'color:{}'.format(color)
growthRate_trafficOnLoanIncluded.style.applymap(color_elements)




# 想来想去，还是直接粘贴比较方便，搞字典或者df再操作来操作去，更麻烦。
objectiveSheet = monthlyOperationReportData.sheets['多头表的数据']
objectiveSheet.range('C3').options(transpose=True).value=list(multiBorrTrafficByCategory_currPeriod[('借款笔数','截至月末数值')])
objectiveSheet.range('E3').options(transpose=True).value=list(multiBorrTrafficByCategory_currPeriod[('借款人数','截至月末数值')])
objectiveSheet.range('G3').options(transpose=True).value=list(multiBorrTrafficByCategory_currPeriod[('余额','截至月末数值')])
objectiveSheet.range('H3').options(transpose=True).value=list(multiBorrTrafficByCategory_lastPeriod[('余额','截至月末数值')])
# 粘贴这个动作可以等数据正确性审阅完成后，将不同数据来源，统一粘贴在多头表的数据sheet中。
# 条件格式中没有变动过大的值就可以复制到Excel。
# 复制的代码也可以先等等，先写COM搞定查询量统计csv下载的代码。





# 读取4-1，感觉也有4-3两个月份的报表格式不一致的问题（数据起始列不一致），也得用4-3那里的代码打印2个4-1的起始列，
# 不一致的话手动修改一致再读取。
multiBorrTrafficByCategory_xlrd = xlrd.open_workbook(r'D:\hlwjrxh\...\2021年12月'+
                                         r'\4-1(2022-01-29)12.xls',
                                         formatting_info=True)
sheetNames = multiBorrTrafficByCategory_xlrd.sheet_names()
print(sheetNames)
objectiveSheet = sheetNames[0]
nRows, nCols = multiBorrTrafficByCategory_xlrd.sheet_by_name(
    objectiveSheet).nrows, multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).ncols
# nrows属性，与xlwings中的current_region不一样，nrows应该是直接取了最右下的单元格计算的，不是取的有数据的最大矩形，而是
# 把左上方的完全空白的列也囊括在内了。
colNames = []
colNames_Row1,colNames_Row2 = [],[]
multiBorrTrafficByCategory_xlrd_dict = {}
for j_row in range(0,nRows):
    if any(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).row_values(j_row)):
        print(j_row)
        break
for j_col in range(0,nCols):
    if any(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(j_col)):
        print(j_col)
        break
for i_col in range(j_col, nCols):
    colNames.append(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row])
    multiBorrTrafficByCategory_xlrd_dict[multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row]] = \
        multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row+1:]
        # 为了直接将excel raw data读入，生成装有raw data的df，不管被merge的表头。
        # 0217但这样不行，因为第一行的合并单元格中，有''作为列名，会重复，导致引用键值为''的键时，python会返回所有键值为''的键，
        # 导致赋值的时候这些键统一赋了值，比如先统一赋了有效D的值，下一个i就会赋有效I，以此类推，for循环走完的时候，所有键值为''的value
        # 都被赋予了最后一个月份的有效G的值。导致dict中键值数小于colNames长度，所以出现很多列都是有效+G的情况。
        # 还是得用read_excel去把业务数据这一部分作为一个multiIndex单独读取出来，再concat基本信息。

    colNames_Row1.append(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row])
    colNames_Row2.append(multiBorrTrafficByCategory_xlrd.sheet_by_name(objectiveSheet).col_values(i_col)[j_row+1])
    
data = pd.DataFrame(multiBorrTrafficByCategory_xlrd_dict, columns=colNames)
# 20210303生成'业务种类信息（全部和有效）也会出现''作为列名的问题。




multiBorrTraffic_currPeriod = pd.read_excel(r'D:\hlwjrxh\...\2021年12月'+
                                         r'\4-1(2022-01-29)12.xls',
                sheet_name=0,header=[j_row,j_row+1], names=None, index_col=[j_col], 
                usecols=None, squeeze=False, dtype={'金额逾期率':np.float64,'项目逾期率':np.float64}, engine=None, 
                converters={'金额逾期率':float,'项目逾期率':float}, true_values=None, false_values=None, skiprows=None, nrows=None, 
                na_values=None, keep_default_na=True, verbose=False, 
                parse_dates=False, date_parser=None, thousands=None, comment=None, 
                skipfooter=0, convert_float=True, mangle_dupe_cols=True)
# skiprows是Rows to skip at the beginning (0-indexed).跳过开头几行。skipfooter是跳过最后的。
multiBorrTraffic_lastPeriod = pd.read_excel(r'D:\hlwjrxh\...\2021年12月'+
                                         r'\4-1(2022-01-29)11.xls',
                sheet_name=0,header=[j_row,j_row+1], names=None, index_col=[j_col], 
                usecols=None, squeeze=False, dtype={'金额逾期率':np.float64,'项目逾期率':np.float64}, engine=None, 
                converters={'金额逾期率':float,'项目逾期率':float}, true_values=None, false_values=None, skiprows=None, nrows=None, 
                na_values=None, keep_default_na=True, verbose=False, 
                parse_dates=False, date_parser=None, thousands=None, comment=None, 
                skipfooter=0, convert_float=True, mangle_dupe_cols=True)
# 如果在use_col中限定了具体的哪些列，则提示cannot specify usecols when specifying a multi-index header。可以先全读了，再slice。
# https://github.com/pandas-dev/pandas/issues/25449，解释了这个问题为什么一直没人处理，人力不够。
# 直到202103也没人处理。只能读取后丢弃了。把不区分有效无效的统计列drop掉。

multiBorrTraffic_currPeriod.drop(columns='比上月增减',level=1,inplace=True)
# multiBorrTrafficByCategory_currPeriod.set_index(keys=multiBorrTrafficByCategory_currPeriod.columns[0:2],inplace=True)
# 老是提示TypeError: unsupported operand type(s) for -: 'str' and 'str'，逾期率有问题，用dtypes、converters都不行，原因在于
# 百分号，需要把它strip掉。用str属性不行，str只能应用于series和index，multiIndex loc取出来的2列形成df，不适用。
# apply(str.strip)不适用，传递过来的变量是series，没有strip方法，applymap只能应用接受一个参数的函数，且这个参数必须是df elements，
# 与strip的参数类型不符。只能替换了。
# 20210305这里不应该用np.nan，应该用''，因为to_replace指代的是符合regex的string，将这个string替换为value，
# 替换了也是object，只能astype({'金额逾期率':float,'项目逾期率':float})不管用，to_numeric就可以，我记得好几次了都是这样的。
# 如果取有特殊性质（如数据类型不同）的列，可以lambda。
multiBorrTraffic_currPeriod=multiBorrTraffic_currPeriod.apply(
    lambda x: pd.to_numeric(x.str.strip('%')) if x.dtype == "object" else x)
# multiBorrTrafficByCategory_currPeriod.index.names=['category','NPL']

multiBorrTraffic_lastPeriod.drop(columns='比上月增减',level=1,inplace=True)
# multiBorrTrafficByCategory_lastPeriod.set_index(keys=multiBorrTrafficByCategory_lastPeriod.columns[0:2],inplace=True)
multiBorrTraffic_lastPeriod=multiBorrTraffic_lastPeriod.apply(
    lambda x: pd.to_numeric(x.str.strip('%')) if x.dtype == "object" else x)
# multiBorrTrafficByCategory_lastPeriod.index.names=['category','NPL']

# 小数点位数太多，搞个条件格式监测数据变化。
growthRate_4_1=(multiBorrTraffic_currPeriod
            -multiBorrTraffic_lastPeriod)/multiBorrTraffic_lastPeriod
# 条件格式，见pandas documentation CH23 Styling章节。
# 代码从网上复制的，网上例子易懂。
def color_elements(val):
    if abs(val) <= 0.1:
        color = 'black'
    elif abs(val)<=0.2:
        color = 'blue'
    else:
        color ='red'
    return 'color:{}'.format(color)
growthRate_4_1.style.applymap(color_elements)




# 4-1粘贴到运行月报数据。
objectiveSheet = monthlyOperationReportData.sheets['多头表的数据']
objectiveSheet.range('K25').options(transpose=True).value=list(multiBorrTraffic_currPeriod[('借款笔数','截至月末数值')][1:-1])
objectiveSheet.range('M25').options(transpose=True).value=list(multiBorrTraffic_currPeriod[('借款人数','截至月末数值')][1:-1])
objectiveSheet.range('O25').options(transpose=True).value=list(multiBorrTraffic_currPeriod[('授信额度','截至月末数值')][1:-1])
objectiveSheet.range('U25').options(transpose=True).value=list(multiBorrTraffic_currPeriod[('余额','截至月末数值')][1:-1])
objectiveSheet.range('W25').options(transpose=True).value=list(multiBorrTraffic_currPeriod[('逾期余额','截至月末数值')][1:-1])
# 粘贴这个动作可以等数据正确性审阅完成后，将不同数据来源，统一粘贴在多头表的数据sheet中。
# 条件格式中没有变动过大的值就可以复制到Excel。
# 小数位数可以用df.round，但这里设置Excel格式更方便。
